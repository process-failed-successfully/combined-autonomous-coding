import sys
import argparse
from pathlib import Path
from typing import List, Dict, Any, Optional, Union
import json
import csv

try:
    import openpyxl
except ImportError:
    openpyxl = None

class ExcelLabManager:
    """
    Manages Excel operations including reading, writing, and inspecting .xlsx files.
    """

    def __init__(self, project_dir: Optional[Path] = None):
        if not openpyxl:
            raise ImportError("openpyxl is required for Excel Lab. Please install it with 'pip install openpyxl'.")
        self.project_dir = project_dir or Path(".")

    def get_info(self, file_path: Union[str, Path]) -> Dict[str, Any]:
        """Returns metadata from an Excel file."""
        file_path = Path(file_path).resolve()
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            sheets = workbook.sheetnames

            return {
                "file": str(file_path),
                "sheets": sheets,
                "sheet_count": len(sheets)
            }
        except Exception as e:
            raise ValueError(f"Error reading Excel file {file_path}: {e}")
        finally:
            if 'workbook' in locals():
                workbook.close()

    def read_sheet(self, file_path: Union[str, Path], sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
        """Reads a specific sheet into a list of dictionaries."""
        file_path = Path(file_path).resolve()
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

            if sheet_name:
                if sheet_name not in workbook.sheetnames:
                     raise ValueError(f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}")
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active

            data = []
            headers = []

            # Iterate rows
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(cell) if cell is not None else f"col_{j}" for j, cell in enumerate(row)]
                    continue

                # Create dict from row
                row_data = {}
                for j, cell in enumerate(row):
                    if j < len(headers):
                        row_data[headers[j]] = cell
                data.append(row_data)

            return data
        except Exception as e:
            raise ValueError(f"Error reading sheet from {file_path}: {e}")
        finally:
            if 'workbook' in locals():
                workbook.close()

def run_excel_lab_logic(args):
    """CLI logic for Excel Lab."""
    try:
        manager = ExcelLabManager(getattr(args, 'project_dir', None))
    except ImportError as e:
        print(f"Error: {e}", file=sys.stderr)
        return False

    if args.action == "info":
        try:
            info = manager.get_info(args.file)
            print(f"--- Excel File Info: {info['file']} ---")
            print(f"Sheet Count: {info['sheet_count']}")
            print("Sheets:")
            for s in info['sheets']:
                print(f"  - {s}")
        except Exception as e:
            print(f"❌ Error: {e}", file=sys.stderr)
            sys.exit(1)

    elif args.action == "read":
        try:
            data = manager.read_sheet(args.file, args.sheet)

            output_format = args.format or "table"

            # Handle output argument logic here if it's meant to save to file
            if args.output:
                output_path = Path(args.output)

                # Determine format from extension if not explicitly set (or default table)
                if output_format == "table" and output_path.suffix == ".csv":
                    output_format = "csv"
                elif output_format == "table" and output_path.suffix == ".json":
                    output_format = "json"

                if output_format == "csv":
                    with open(output_path, 'w', newline='', encoding='utf-8') as f:
                        writer = csv.DictWriter(f, fieldnames=data[0].keys() if data else [])
                        writer.writeheader()
                        writer.writerows(data)
                    print(f"✅ Saved to {output_path}")
                elif output_format == "json":
                     with open(output_path, 'w', encoding='utf-8') as f:
                        def json_serial(obj):
                            if hasattr(obj, 'isoformat'):
                                return obj.isoformat()
                            return str(obj)
                        json.dump(data, f, indent=2, default=json_serial)
                     print(f"✅ Saved to {output_path}")
                else:
                    # Fallback or error? For now, if output is specified but format is table/unknown
                    # we might just error or try to print to file?
                    # Let's assume user wants to print table to stdout if format is table
                    pass

            # Output to stdout if no file output (or if table format requested regardless of file)
            if not args.output or output_format == "table":
                if output_format == "table":
                    try:
                        from rich.console import Console
                        from rich.table import Table
                        console = Console()
                        table = Table(show_header=True, header_style="bold magenta")

                        if not data:
                            print("Sheet is empty.")
                            sys.exit(0)

                        headers = list(data[0].keys())
                        for h in headers:
                            table.add_column(h)

                        limit = getattr(args, 'limit', 50)
                        for row in data[:limit]:
                            table.add_row(*[str(row.get(h, "")) for h in headers])

                        console.print(table)
                        if len(data) > limit:
                            console.print(f"[dim]Showing first {limit} of {len(data)} rows.[/dim]")
                    except ImportError:
                        # Fallback to CSV-like print
                        if not data:
                            print("Sheet is empty.")
                            sys.exit(0)
                        headers = list(data[0].keys())
                        print(",".join(headers))
                        for row in data:
                            print(",".join(str(row.get(h, "")) for h in headers))

                elif output_format == "csv":
                    writer = csv.DictWriter(sys.stdout, fieldnames=data[0].keys() if data else [])
                    writer.writeheader()
                    writer.writerows(data)

                elif output_format == "json":
                    # Handle non-serializable objects (like datetime)
                    def json_serial(obj):
                        if hasattr(obj, 'isoformat'):
                            return obj.isoformat()
                        return str(obj)
                    print(json.dumps(data, indent=2, default=json_serial))

        except Exception as e:
            print(f"❌ Error reading Excel file: {e}", file=sys.stderr)
            sys.exit(1)

    return True
